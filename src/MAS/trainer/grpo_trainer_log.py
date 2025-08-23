import os
import re
import pandas as pd
from trl import GRPOTrainer
from typing import Optional
from openpyxl import Workbook, load_workbook
from MAS.utils.logging_utils import setup_logger

SAVE_PATH = "grpo_trainer_log"

_ILLEGAL_CHAR_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

logger = setup_logger()

def _clean_cell(val):
    if isinstance(val, str):
        return _ILLEGAL_CHAR_RE.sub('', val)
    return val


def save_to_xlsx(data, xlsx_path):
    flattened_data = {}
    for key, value in data.items():
        if isinstance(value, dict):
            for nested_key, nested_value in value.items():
                flattened_data[nested_key] = nested_value
        else:
            flattened_data[key] = value

    df = pd.DataFrame(flattened_data).map(_clean_cell)

    file_exists = os.path.exists(xlsx_path)
    if file_exists:
        try:
            existing_df = pd.read_excel(xlsx_path)
            startrow = existing_df.shape[0] + 1
        except Exception:
            startrow = 0
    else:
        os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
        wb = Workbook()
        wb.save(xlsx_path)
        startrow = 0

    with pd.ExcelWriter(xlsx_path, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
        sheet_name = writer.book.sheetnames[0]
        df.to_excel(writer, index=False, header=(startrow == 0), startrow=startrow, sheet_name=sheet_name)

    wb = load_workbook(xlsx_path)
    ws = wb.active
    for column_cells in ws.columns:
        key_length = len(str(column_cells[0].value))
        value_length = max([len(str(cell.value)) for cell in column_cells[1:] if cell.value is not None], default=0)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(key_length*2, value_length+5), 60)

    wb.save(xlsx_path)
    return df


class GRPOTrainerWithLog(GRPOTrainer):
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._last_xlsx_saved_step = -1
    
    def log(self, logs: dict[str, float], start_time: Optional[float] = None) -> None:
        super().log(logs, start_time)
        
        if self.accelerator.is_main_process and self.log_completions:
            mode = "eval" if self.control.should_evaluate else "train"
            if self._last_xlsx_saved_step != self._last_loaded_step:
                self._last_xlsx_saved_step = self.state.global_step
                table = {
                    "step": [str(self.state.global_step)] * len(self._textual_logs["prompt"]),
                    "prompt": self._textual_logs["prompt"],
                    "completion": self._textual_logs["completion"],
                    **self._textual_logs["rewards"],
                }

                excel_path = os.path.join(self.args.output_dir, f"{SAVE_PATH}_{mode}.xlsx")
                
                try:
                    save_to_xlsx(table, excel_path)
                except Exception as e:
                    logger.warning(f"Warning: Failed to save log to Excel: {e}")